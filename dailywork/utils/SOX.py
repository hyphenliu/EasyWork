#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@File  : SOX.py
@Author: HP.Liew
@Date  : 2019/8/22 16:36
@Desc  : 
'''
import os
import datetime
from django.core.cache import cache
from openpyxl import Workbook, load_workbook
from EasyWork.settings import DOWNLOAD_DIRS
from EasyWork.utils.database_ops import *

sox_list = ['stand_point', 'province_point', 'area', 'procedure', 'sub_procedure', 'control_goal',
            'standard_describe', 'company_describe', 'frequency', 'control_type', 'control_method', 'department_list',
            'duty', 'classification', 'reference_file', 'focus_point', 'test_file', 'action']
sox_feature = ['标准控制点编号', '控制点编号', '适用范围', '业务流程', '子流程', '控制目标', '标准控制点描述', '控制点描述', '频率', '控制类型', '控制方式',
               '部门', '负责人', '控制点分类', '参考文件', '建议的关注点', '穿行测试资料', '材料提供']


def getXlsContent(tableName, filename, department='基础平台'):
    '''
    主程序入口
    提取部门控制矩阵
    写入原始表单数据{'sheetname1':[[col1,col2,...],...],...}
    :param filename:
    :return: {'省控制点编号':{'字段':字段值,...}}
    '''
    ws_result = {}
    warning_msgs = []
    ws_ori_result = {}  # 保存原始信息
    staff_dict = _getContactList(department)  # {name/(IT): phone}
    # wb = load_workbook(filename, read_only=True, data_only=True)
    wb = load_workbook(filename)
    for sn in wb.sheetnames:  # 处理每个sheet

        print('sheet name:{0}'.format(sn))
        # ws = wb[sn]
        ws = _unmergeCell(wb[sn])
        feature_cols = _getHeadLine(ws)
        if not feature_cols: continue
        ws_ori_result[sn] = []  # 保留原始表格sheet名
        nrows = ws.max_row  # 最大行数
        ncols = ws.max_column  # 最大列数
        start_row_num = feature_cols['row_num']
        feature_cols.pop('row_num')
        # 保存表头信息
        head_line = [_cellValue(c) for c in ws[start_row_num]]
        head_line.insert(1, '部门责任人')
        ws_ori_result[sn].append(head_line)
        # 处理每一行
        for row in range(start_row_num, nrows + 1):
            item = {}
            # 提取部门员工，未含本部门员工则跳过
            staff_cell = ws.cell(row=row, column=feature_cols['duty']).value
            if not staff_cell: continue
            # 提取责任人
            staff_list = _extractContactor(staff_cell, staff_dict)
            if not staff_list: continue
            department_cell = ws.cell(row=row, column=feature_cols['department_list']).value
            for k, v in feature_cols.items():
                if v < 0:
                    item[k] = '未提取到'
                    continue
                item[k] = _cellValue(ws.cell(row=row, column=v))
            # 提取可能存在问题的信息
            if not _departmentFilter(department_cell, staff_cell):
                warning_msgs.append(item['stand_point'] + ": " + staff_list)
                continue
            item['duty'] = staff_cell
            item['staff'] = staff_list
            if item['province_point'].strip():
                point = item['province_point']
            else:
                point = item['stand_point']
            if point and not point in ws_result:  # 不重复控制点
                # 提取本门控制点原始信息
                o_list = [_cellValue(c) for c in ws[row]]
                o_list.insert(1, staff_list)  # 插入部门联系人信息
                ws_ori_result[sn].append(o_list)
                ws_result[point] = item
    # 写入原始数据到文件中，并在redis中提供下载刚刚生成的文件
    _writeContent(tableName, filename, ws_ori_result, department)
    if warning_msgs:
        print(warning_msgs)
        warning = '<div style="background:#FF0"><p>控制点存在责任人部门信息异常：</p>{}</div>'.format(''.join(['<p>{}</p>'.format(i) for i in warning_msgs]))
        return ws_result, warning
    return ws_result, ''


def _unmergeCell(sheet_content):
    '''
    拆分并填充合并单元格
    :param sheet_content:
    :return:
    '''
    m_cells = sheet_content.merged_cells.ranges
    m_cells_dict = {}
    for mc in m_cells:
        cell_value = sheet_content.cell(row=mc.min_row, column=mc.min_col).value
        m_cells_dict[mc.coord] = [mc, cell_value]
    for k, v in m_cells_dict.items():
        merge_all_list = []
        sheet_content.unmerge_cells(k)
        r1, r2, c1, c2 = v[0].min_row, v[0].max_row, v[0].min_col, v[0].max_col
        if (r1 != r2 and c1 != c2):
            row_col = [(x, y) for x in range(r1, r2 + 1) for y in range(c1, c2 + 1)]
            merge_all_list.extend(row_col)
        elif (r1 == r2 and c1 != c2):  # or (r1 != r2 and c1 == c2):
            col = [(r1, n) for n in range(c1, c2 + 1)]
            merge_all_list.extend(col)
        elif (r1 != r2 and c1 == c2):
            row = [(m, c1) for m in range(r1, r2 + 1)]
            merge_all_list.extend(row)
        for mal in merge_all_list:
            sheet_content.cell(row=mal[0], column=mal[1]).value = v[1]

    return sheet_content


def _cellValue(cell):
    if not cell.value:
        return ' '
    if isinstance(cell.value, str):
        return cell.value.replace('\n', '')
    if isinstance(cell.value, datetime.datetime):
        return cell.value.strftime('%Y-%m-%d')
    if isinstance(cell.value, int):
        return str(cell.value)
    return cell.value


def _getHeadLine(sheet_content):
    '''
    返回字段所在的行/列号，没有找到的字段列号为-1
    :param sheet_content:
    :return: result{sox_list:column_num}
    '''
    result = {}
    col_num = []
    department_flag = False  # 区分“具体部门”和“部门”
    for sl in sox_list:
        result[sl] = -1
    row_num = -1
    # 找到首行
    for row in sheet_content.rows:
        count = 0
        cell_values = '|'.join([_cellValue(c) for c in row])
        # 删除换行符
        # cell_values = cell_values.replace('\n', '').strip('|')
        cell_values = re.sub(r'\s', '', cell_values).strip('|')
        for sf in sox_feature:
            if sf in cell_values: count += 1
        if count > len(sox_feature) / 4:
            row_num = row[1].row
            if '具体部门' in cell_values: department_flag = True
            break
    if row_num < 0:
        print('【ERROR】解析Excel出错，没有找到超过1/4的特征值列')
        return
    # 找到字段对应的列
    result['row_num'] = row_num
    for cell in sheet_content[row_num]:
        cv = cell.value
        if not isinstance(cv, str) or not cv:
            # print('【ERROR】解析Excel第{0}行出错，存在单元格为空'.format(row_num))
            continue
        for sf in sox_feature:
            key = sox_list[sox_feature.index(sf)]
            if sf == '部门' and department_flag: sf = '具体部门'
            if sf in cv and (result[key] < 0) and (cell.column not in col_num):
                result[key] = cell.column
                col_num.append(cell.column)
    return result


def _extractContactor(xls_data, staff_dict):
    '''
    提取出部门控制点责任人
    :param xls_data:
    :param staff_dict:# {name/(IT): phone}
    :return: '责任人1,责任人2,...
    '''
    result = []
    xls_data = xls_data.replace('基础平台部','').replace('基平','')
    pattern = re.compile('[^\u4e00-\u9fa5]')
    data = pattern.split(xls_data)  # 提取汉字
    index_cn = [xls_data.index(d) for d in data if d]  # 记录其他信息
    index_cn.append(len(xls_data))
    # 切分字段，方便对比手机号
    xls_data_items = [xls_data[index_cn[i]:index_cn[i + 1]] for i in range(len(index_cn) - 1)]
    for xdi in xls_data_items:
        if len(xdi) < 2: continue
        items = pattern.split(xdi) # 提取中文，含名字
        if re.findall(r'\d{11}', xdi):
            phone_flag = True  # 是否包含手机号码
        else:
            phone_flag = False
        for k, v in staff_dict.items():
            k = pattern.split(k)[0]  # 去除“（IT）”字样
            if k in items:
                if phone_flag and not v in xdi:
                    continue
                result.append(k)
    return ','.join(result)


def _departmentFilter(dep_data, staff_cell):
    # for i in ['基础','基平', '实物管理部门', '工程建设部门', '各采购需求部门', '各采购验收部门', '实施部门']:
    for i in ['基础', '基平', '实物管理', '工程建设', '各', '实施']:
        if i in dep_data:
            return True
    if '基础' in staff_cell:
        return True


def _getContactList(department):
    '''
    根据部门查询部门员工姓名和电话
    :param department:
    :return: {'姓名':电话,...}
    '''
    result = {}
    datas = getSingleData('contact', 'department', department).values()
    for data in datas:
        result[data['name']] = data['phone']
    return result


def _writeContent(tableName, filename, data, department):
    file_pre, ext = os.path.splitext(filename)
    filename = '{0}-{1}{2}{3}'.format(file_pre[:-20], department, file_pre[-20:-9], ext)  # 去掉自动添加上去的时间戳，添加部门
    file_item = filename.split(os.sep)
    file_item[-2] = 'download'
    filename = os.sep.join((file_item))
    if os.path.exists(filename): os.remove(filename)  # 删除已经存在的文件
    wb = Workbook()
    for k, v in data.items():
        ws = wb.create_sheet(k, 0)
        for row in range(len(v)):
            for col in range(len(v[row])):
                ws.cell(row=row + 1, column=col + 1, value=v[row][col])  # cell 起始位置必须是1而非0
    wb.save(filename)
    wb.close()
    cache.set('download{0}{1}file'.format('dailywork', tableName), filename, 1 * 60)
