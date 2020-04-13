#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@File  : mail_sender.py
@Author: HP.Liew
@Date  : 2019/5/7 10:29
@Desc  :
'''
import ssl
import random
import logging
import shutil
import openpyxl
from selenium import webdriver

from EasyWork.utils.passwd_ops import *
from EasyWork.utils.mail_utils import *
from EasyWork.utils.config import Config

ssl._create_default_https_context = ssl._create_unverified_context
logger = logging.getLogger('console')


class CheckDevice:
    def _initVar(self):
        '''
        初始化时间数值和其他相关参数，只能运行一次初始化一次，不然就会有问题！
        :return:
        '''
        self.configer = Config('devicecheck.ini')
        self.department = self.configer.get('device_check', 'department')
        self.inspector = self.configer.get('device_check', 'inspector')
        self.checker = self.configer.get('device_check', 'checker')
        self.sender = self.configer.get('device_check', 'sender')
        self.smtp_server = self.configer.get('device_check', 'mailserver')
        self.passwd = getPassword('device_check_mailpasswd')
        self.mail_address = splitFormatEmailAddr(self.sender)
        #############################################################################
        ct = time.localtime(time.time())
        year, month, day, hour, minute = ct.tm_year, ct.tm_mon, ct.tm_mday, ct.tm_hour, ct.tm_min
        hour, day = (24, day - 1) if hour == 0 else (hour, day)
        self.hourStr = f"{hour - 2:0>2d}点-{hour:0>2d}点"  # 06:00-08:00
        self.dateStr1 = f'{month:0>2d}{day:0>2d}'  # 0615
        self.dateStr2 = f'{year}年{month:0>2d}月{day:0>2d}日'  # 2019年06月15日
        self.dateStr3 = f'{month:0>2d}月{day:0>2d}日{self.hourStr}'  # 06月15日18:00点
        self.reportFileName = os.path.join(settings.TEST_DIRS, 'doc',
                                           '核心设备每日检查表-深圳{}({}).xlsx'.format(self.dateStr1, self.hourStr))
        self.reportTempFile = os.path.join(settings.TEST_DIRS, 'doc', '核心设备每日检查表-深圳-模版.xlsx')
        self.zip_file = os.path.join(settings.TEST_DIRS, 'doc', f'核心设备每日检查表-深圳{self.dateStr1}.zip')
        self.zip_base_name = os.path.split(self.zip_file)[-1]
        if os.path.exists(self.zip_file):
            os.remove(self.zip_file)

    def _initBroswer(self):
        '''
        初始化无头浏览器
        :return:
        '''
        options = webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--headless')
        self.browser = webdriver.Chrome(executable_path=settings.CHROME_DRIVER, chrome_options=options)
        self.browser.implicitly_wait(20)

    def _check(self):
        '''
        设备巡检
        :return: 返回巡检失败的主机
        '''
        failedList = []
        try:
            self._initBroswer()
        except Exception as e:
            print(f'[ERROR] Init browser failed.{e}')
            self.browser.quit()
            return ['failed', 'failed']

        for host in ['pa', 'f5', 'sw']:
            ips = re.split('[,，]', self.configer.get(host, 'ip'))
            keywords = self.configer.get(host, 'keywords')
            userid = self.configer.get(host, 'userid')
            user = self.configer.get(host, 'user')
            passwordid = self.configer.get(host, 'passwordid')
            password = getPassword(f'device_check_{host}')
            submit = self.configer.get(host, 'submit')
            infourl = re.split('[,，]', self.configer.get(host, 'infourl'))
            for ip in ips:
                print('*' * 79)
                print(f'[INFO] Start checking [{ip}] ...')
                logger.info(f'Start checking [{ip}] ...')

                self.browser.get(f'https://{ip.strip()}')
                if self.browser.page_source == '<html><head></head><body></body></html>':
                    print(f'[ERROR] Connect {ip} failed. Network error.')
                    logger.error(f'Connect {ip} failed. Network error.')
                    failedList.append(f'Connect {ip} failed')
                    continue
                self.browser.find_element_by_id(userid).send_keys(user)
                self.browser.find_element_by_id(passwordid).send_keys(password)
                self.browser.find_element_by_css_selector(submit).click()
                content = str(self.browser.page_source)

                if keywords in ' '.join(content.split('\n')):
                    print(f'[INFO] Login {ip} Success')
                    logger.info(f'Login {ip} Success')
                else:
                    print(f'[ERROR] Login {ip} Failed')
                    logger.error(f'{ip} login failed')
                    failedList.append(f'{ip} login failed')

                for url in infourl:
                    print(f'[INFO] Visiting {url}')
                    logger.info(f'Visiting {url}')
                    int = random.randint(10, 15)
                    time.sleep(int)
                    self.browser.get(f'https://{ip}/{url.strip()}')
                print(f'[INFO] Logout {ip} .')
                logger.info(f'Logout {ip} .')
        print('[INFO] Check complete!')
        logger.info('Check complete!')
        self.browser.quit()
        return failedList

    def _report(self):
        '''
        生成报告
        :return:返回是否生成报告成功
        '''
        print('[INFO] Product report')
        shutil.copy(self.reportTempFile, self.reportFileName)
        wb = openpyxl.load_workbook(self.reportFileName)
        ws = wb.worksheets[0]
        ws.title = self.dateStr1
        ws["C1"] = f"检查时段{self.dateStr3}"
        for i in range(3, 12):
            ws.cell(i, 5, self.inspector)
            ws.cell(i, 6, self.checker)
        try:
            wb.save(self.reportFileName)
            print('[INFO] Product report {} success!'.format(self.reportFileName))
            return True
        except Exception as e:
            print('[ERROR] Product report {} failed. {}'.format(self.reportFileName, e))
            return False

    def _compress(self):
        '''
        加密压缩报告
        :return: 返回是否压缩成功
        '''
        print('[INFO] Compress cipher zip file')
        compress_key = getPassword('device_check_compress')
        if settings.op_system == 'Windows':
            compress_cmd = 'Bandizip.exe a -y  -p:{} {} {}'.format(compress_key, self.zip_file, self.reportFileName)
        else:
            compress_cmd = 'zip -rP:{} {} {}'.format(compress_key, self.zip_file, self.reportFileName)
        time.sleep(random.randint(10, 15))
        try:
            os.system(compress_cmd)
            print('[INFO] Compress zip {} success!'.format(self.zip_file))
            return True
        except Exception as e:
            print('[ERROR] Compress zip {} encounter en error!'.format(self.zip_file))
            print(e)
            return False

    def _send(self):
        '''
        发送邮件
        :return: 返回是否发送成功
        '''
        recevier = re.split('[,，]', self.configer.get('device_check', 'recevier'))
        recevier = getMulContactEmailAddr(names=recevier, department=self.department)
        cc = re.split('[,，]', self.configer.get('device_check', 'cc'))
        cc = getMulContactEmailAddr(names=cc, department=self.department)
        mailsign = self.configer.get('device_check', 'mailsign')

        print('[INFO] Prepare send email')
        title = '{}{}深圳侧核心设备巡检'.format(self.dateStr2, self.hourStr)
        body = """各位好，
            {}{}深圳核心设备巡检无异常，检查人：{}、复核人：{}，详见附件{}""".format(
            self.dateStr2, self.hourStr, self.inspector, self.checker, mailsign)
        return mailSender(title=title, sender=self.sender, recevier=recevier, cc=cc, body=body,
                          sender_mail=self.mail_address, sender_pass=self.passwd, file_name=self.zip_file)

    def _program(self):
        self._initVar()
        # time.sleep(random.randint(0, 1) * 60)
        result = self._check()
        if result:  # 巡检失败
            return result
        time.sleep(random.randint(0, 3) * 60)
        if not self._report():  # 生成报告失败
            return 'Gener report failed.'
        time.sleep(random.randint(0, 3) * 60)
        if not self._compress():  # 压缩文件失败
            return 'Compress cipher zip file failed'
        time.sleep(random.randint(0, 3) * 60)
        if not self._send():  # 发送邮件失败
            return 'Send email encounter error'
        return 'success'

    def deviceCheck(self):
        '''
        巡检和
        :return:
        '''
        for i in range(3):
            if not self._program() == 'success':
                time.sleep(1 * 60)
            else:
                break
