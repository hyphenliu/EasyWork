import os
from ipaddress import *
import re
import socket
from datetime import datetime, date
from collections import defaultdict, Counter
from django.core.cache import cache
from openpyxl import Workbook, load_workbook
from EasyWork.utils.database_ops import *
from EasyWork.utils.data_struct import *

access_list = ['direction', 'source_IP', 'source_map_IP', 'source_port', 'dest_IP', 'dest_map_IP', 'dest_port',
               'transport_protocal', 'app_protocal', 'access_use', 'vpn_domain']
access_feature = ['访问方向', '源地址', '源地址对应的映射IP', '源端口', '目的地址', '目的地址对应的映射IP', '目的端口', '传输层协议', '应用层协议', '策略用途',
                  '接入IP承载网所属VPN域']
province_list = ['北京', '天津', '上海', '重庆', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽', '福建', '江西', '山东', '河南', '湖北',
                 '湖南', '广东', '海南', '四川', '贵州', '云南', '陕西', '甘肃', '青海', '台湾', '内蒙', '广西', '西藏', '宁夏', '新疆', '香港', '澳门']
hq_list = ['信息港', '南方基地', '深圳', 'ucenter', '集团', '哈池', '呼池']


def _isIp(address):
    address = address.strip()
    try:
        ip_address(address)
        return True
    except Exception as e:
        return False


def _cellValue(cell):
    if not cell.value:
        return ''
    if isinstance(cell.value, str):
        return cell.value.replace('\n', ' ').strip()
    if isinstance(cell.value, int):
        return str(cell.value)
    if isinstance(cell.value, datetime):
        return cell.value.strftime('%Y%m%d')

    return cell.value


def _getHeadLine(sheet_content):
    '''
    返回字段所在的行/列号，没有找到的字段列号为-1
    :param sheet_content:
    :return: result{access_list_item:column_num}
    '''

    result = {}
    col_nums = []
    warning = ''
    for sl in access_list:
        result[sl] = -1
    row_num = -1
    m_cells = sheet_content.merged_cells  # 获取所有的合并单元格，以供后面判断
    # 找到首行
    for row in sheet_content.rows:
        count = 0
        cell_values1 = '|'.join([_cellValue(c) for c in row])
        cell_values = cell_values1.replace('\n', '').strip(' |')
        if (len(cell_values1) - len(cell_values)) > 20:
            warning = '表格[{}]存在大量无效空列，影响程序运行速度； '.format(sheet_content.title)
            print(warning)
        # result['ncols'] = len(cell_values.split('|')) # 有效列数，有的表格可能有好多空列，影响到运行速度
        for af in access_feature:
            if af in cell_values: count += 1
        if count > len(access_feature) / 2:
            row_num = row[1].row
            break
    if row_num < 0:
        print('[ERROR] 解析Excel出错，表单【{0}】没有找到超过半数的特征值行; '.format(sheet_content.title))
        return {}, warning
    print('[SUCCESS] 找到有效表头的表单【{}】'.format(sheet_content.title))
    # 找到字段对应的列
    result['row_num'] = row_num
    for cell in sheet_content[row_num]:
        print(cell.column)
        print(cell.row)
        cv = _cellValue(cell)
        cv = re.sub(r'\s', '', cv)
        if not isinstance(cv, str) or not cv:
            # print('【ERROR】解析Excel第{0}行出错，存在单元格为空'.format(row_num))
            continue
        for af in access_feature:
            key = access_list[access_feature.index(af)]  # 根据特征值找到对应的存储字典key值
            if af in cv and (result[key] < 0) and (cell.column not in col_nums):
                if af in ['源端口', '目的端口', '传输层协议']:
                    first_col = sheet_content.cell(row=cell.row + 1, column=cell.column).value
                    if not af == '传输层协议':
                        if first_col == '从':
                            s = cell.column
                            d = cell.column + 1
                        else:
                            s = cell.column + 1
                            d = cell.column
                        if af == '源端口':
                            result['source_port_from'] = s
                            result['source_port_to'] = d
                        else:
                            result['dest_port_from'] = s
                            result['dest_port_to'] = d
                    else:
                        if first_col == 'tcp':
                            t = cell.column
                            u = cell.column + 1
                        else:
                            t = cell.column + 1
                            u = cell.column
                        result['transport_protocal_tcp'] = t
                        result['transport_protocal_udp'] = u
                else:
                    result[key] = cell.column
                col_nums.append(cell.column)
            elif ('映射IP' in af) and ('映射IP' in cv) and (cell.column not in col_nums):
                if '源地址' in cv:
                    result['source_map_IP'] = cell.column
                    col_nums.append(cell.column)
                elif '目的地址' in cv:
                    result['dest_map_IP'] = cell.column
                    col_nums.append(cell.column)
                else:
                    warning += '映射地址表头[{}]解析错误; '.format(cv)
    for i in ['transport_protocal', 'source_port', 'dest_port']:
        result.pop(i)
    unresoled = []
    for k, v in result.items():
        if v < 0:
            unresoled.append(access_feature[access_list.index(k)])
    if unresoled:
        warning += '【{}】解析失败； '.format(', '.join(unresoled))
    print('[WARNING]', warning)
    return result, warning


def readXlsContent(tableName, filename):
    '''
    读取excel表格内容，获取网络策略开通单的表头，提取表头对应列的数据
    :param filename:
    :return:data=[['direction', 'source_IP', 'source_map_IP', 'source_port', 'dest_IP', 'dest_map_IP', 'dest_port',
                    'transport_protocal', 'app_protocal', 'access_use', 'vpn_domain']......],
            msg_dict={'error':'','warning':''} # HTML格式的字符串
    '''
    print('处理文件：{}'.format(filename))
    data = []
    message = ''
    warning_msgs = []
    error_msgs = []
    wb = load_workbook(filename, data_only=True)
    sheet_list = wb.sheetnames
    for sheet_name in sheet_list:
        errors = []  # 存储检查本表单发现的错误信息
        sheet = wb[sheet_name]
        sheet_name = sheet.title
        print('处理表单：{0}'.format(sheet_name))
        head_index, warning = _getHeadLine(sheet)
        if not head_index:
            continue
        # print(head_index)
        nrows = sheet.max_row
        start_row = head_index['row_num'] + 1
        head_index.pop('row_num')
        if nrows < start_row + 1:
            print('{0}表格内容为空，请检查！'.format(sheet.title))
            return
        # 一行一行处理,最后将每一行的所有错误信息汇总
        for row in range(start_row + 1, nrows + 1):
            if not ''.join([_cellValue(c) for c in sheet[row]]): continue
            item_dict = {}  # 记录每一列的信息,生成键值对
            error_msg = []  # 记录每一行的错误信息
            transport_protocal_tcp = sheet.cell(row=row, column=head_index['transport_protocal_tcp']).value
            transport_protocal_udp = sheet.cell(row=row, column=head_index['transport_protocal_udp']).value
            for k, v in head_index.items():
                if v < 0:  # 字段没有找到
                    item_dict[k] = ''
                    continue
                err_msg = ''
                cell_value = _cellValue(sheet.cell(row=row, column=v))  # 提取单元格内容
                if 'IP' in k:
                    IP_list, err_msg = _extractIP(cell_value, access_feature[access_list.index(k)])
                    item_dict[k] = '\n'.join(IP_list)
                elif '_port' in k:
                    port_dict, err_msg = _extractPort(k, cell_value, transport_protocal_tcp, transport_protocal_udp)
                    if isinstance(port_dict, dict):
                        for i, j in port_dict.items():
                            port_dict[i] = '\n'.join(str(n) for n in j)
                        item_dict[k] = port_dict
                    else:
                        item_dict[k] = '\n'.join(str(i) for i in port_dict)
                elif k == 'direction':  # 提取访问方向信息
                    acc_list, err_msg = _extractProvince(cell_value, access_feature[access_list.index(k)])
                    item_dict[k] = acc_list
                else:
                    item_dict[k] = cell_value
                # 去除重复的错误
                if not (err_msg.strip() in error_msg):
                    error_msg.append(err_msg.strip())
            # 汇总错误信息
            error_msg = ''.join(error_msg)
            # 处理每一列的信息，将字典值转化为列表值，处理端口信息
            item_list, err_msg = _sortItem(item_dict)
            error_msg += err_msg
            # 每一行，只要出现错误信息就不添加到最终结果中
            if error_msg.strip():
                error_msg = '第[{0:^{1}d}]行数据错误：{2}'.format(row, len(str(nrows)), error_msg)
                errors.append(error_msg)
                # print(error_msg)
            else:
                for il in item_list:
                    data.append(il)
        if warning:  # 合并告警信息
            warning_msgs.append('表单<b>《{}》</b>存在的告警信息：'.format(sheet_name))
            warning_msgs.append(warning)
        if errors:  # 合并错误信息
            error_msgs.append('表单<b>《{}》</b>存在的错误信息：'.format(sheet_name))
            error_msgs.extend(errors)
    # print(data, {'warning': warning_msgs, 'errors': error_msgs})
    if warning_msgs:
        message += '<div style="background:#FF0">{}</div>'.format(
            ''.join(['<p>{}</p>'.format(i) for i in warning_msgs]))
    if error_msgs:
        message += '<div style="color:#F00">{}</div>'.format(''.join(['<p>{}</p>'.format(i) for i in error_msgs]))
    # 写入原始数据到文件中，并在redis中提供下载刚刚生成的文件
    # _writeContent(tableName, filename, data)
    return data, message


def _extractIP(cell_content, column_name):
    '''
    提取并判断IP地址的正确性
    :param cell_content:
    :return:[ip1,ip2]
    '''

    if not cell_content:
        return [], '<u><b>{}</b></u>单元格为空; '.format(column_name)
    result = []
    error_msg = []
    cell_content = cell_content.lower()
    pattern = re.compile(r'[^\d/.~-]+?')

    items = re.split(pattern, cell_content)
    for item in items:
        if not item.strip(): continue
        if item == '/': continue
        if not re.match(r'^(\d{1,3}\.){3}\d{1,3}', item):
            e_msg = '[{0}]不是IP地址; '.format(item)
            if not e_msg in error_msg:
                error_msg.append(e_msg)
            continue
        if '-' in item or '~' in item:
            ip_list = _completeIPAddress(item)
            if not ip_list:
                e_msg = '[{0}]IP无法解析; '.format(item)
                if not e_msg in error_msg:
                    error_msg.append(e_msg)
                continue
            result.extend(ip_list)
        else:
            result.append(item)
    IP_list = []
    for item in result:
        if not _isIp(item):
            e_msg = '[{0}]IP地址不合法; '.format(item)
            if not e_msg in error_msg:
                error_msg.append(e_msg)
            continue
        if item.startswith('255'):
            e_msg = '[{0}]为子网掩码; '.format(item)
            if not e_msg in error_msg:
                error_msg.append(e_msg)
            continue
        if item in IP_list:
            e_msg = '[{}]地址重复; '.format(item)
            if not e_msg in error_msg:
                error_msg.append(e_msg)
            continue
        IP_list.append(item)
    if not result:
        e_msg = '<u><b>{}</b></u>没有提取到IP地址; '.format(column_name)
        if not e_msg in error_msg:
            error_msg.append(e_msg)
    return sorted(result), ''.join(error_msg)


def _extractProvince(cell_content, column_name):
    '''
    提取策略访问方向
    :param cell_content:
    :return: [源，目的，[中间]]
    '''
    if not cell_content:
        return [], '<u><b>{}</b></u>单元格为空; '.format(column_name)
    if not '>' in cell_content:
        return ['', '', ''], '<u><b>{}</b></u>[{}]填写有误; '.format(column_name, cell_content)
    error_msg = ''
    access_from_checked = False
    access_to_checked = False
    cell_content = cell_content.lower()
    cell_content = re.sub(r'[—-]{0,}>', '>', cell_content)
    items = cell_content.split('>')
    if len(items) < 2:
        return [], '<u><b>{}</b></u>未使用“->”符号区分访问方向; '.format(column_name)
    access_from = items[0]
    access_to = items[-1]
    for pl in province_list:
        if pl in access_from:
            items[0] = pl
            access_from_checked = '省端'
        if pl in access_to:
            items[-1] = pl
            access_to_checked = '省端'
        if access_from_checked and access_to_checked:
            break
    for pl in hq_list:
        if pl in access_from:
            items[0] = pl
            access_from_checked = '总部'
        if pl in access_to:
            items[-1] = pl
            access_to_checked = '总部'
        if access_from_checked and access_to_checked:
            break
    if not access_from_checked:
        error_msg += '源位置信息[{0}]不在省或总部列表中; '.format(access_from)
    if not access_to_checked:
        error_msg += '目的位置信息[{0}]不在省或总部列表中; '.format(access_to)
        print(error_msg)
    # print(access_from_checked, '|'.join(access_middle), access_to_checked)
    if not (access_from_checked and access_to_checked):
        error_msg += '[{0}]提取访问方向错误; '.format(cell_content)
    route = '->'.join(items)
    return route.upper(), error_msg


def _completeNumber(content):
    '''
    补充连续的数字
    :param content:需要拆分的连续数字字符串
    :return:
    '''
    result = []
    items = re.split('[-~]', content)
    if not items:
        return
    for i in range(int(items[0]), int(items[-1]) + 1):
        result.append(i)
    return result


def _completeIPAddress(content):
    '''
    补全连续的IP地址
    :param content:
    :return:
    '''
    if not re.findall('[-~]', content.split('.')[-1]):
        return
    result = []
    items = content.split('.')
    for i in _completeNumber(items[-1]):
        result.append('{}.{}'.format('.'.join(items[:3]), i))
    return result


def _extractNumber(line):
    '''
    # 提取字符串中的数字和连续符号
    :param line: 输入字符串
    :return: 拆分后的端口号
    '''
    error_msg = ''
    result = []
    # 提取端口号中的数字和连续符号
    items = [item for item in re.split('[^\d~-]+?', line) if item]
    if not items:
        return result, error_msg
    for item in items:
        if not re.match('^\d+$', item):
            ports = _completeNumber(item)
            if not ports:
                error_msg += '[{0}]端口解析错误; '.format(item)
                continue
            for p in ports:
                if p > 65535 or p < 1:
                    error_msg += '[{0}]端口非法; '.format(item)
                    continue
            result.extend(ports)
        elif int(item) > 65535 or int(item) < 1:
            error_msg += '[{0}]端口非法; '.format(item)
            continue
        else:
            result.append(int(item))
    return result, error_msg


def _extractPort(k, cell_content, tcp, udp):
    '''
    提取端口号
    :param cell_content:
    :param tcp:
    :param udp:
    :return: {'tcp':[],'udp':[]}
    '''
    error_msg = ''
    # 单元格内容为空
    if not cell_content.strip():
        if k.startswith('source_port'):
            return ['不限'], error_msg
        return [], error_msg
    # 单元格内容不为空
    cell_content = cell_content.lower().replace('\n', ' ')
    if '不限' in cell_content:
        return ['不限'], error_msg
    tcp_ports = []
    udp_ports = []
    if tcp and udp:
        if not ('tcp' in cell_content and 'udp' in cell_content):
            error_msg += '同时选择了TCP和UDP，但是端口{0}中未区分TCP和UDP; '.format(cell_content)
        else:
            tcp_index = cell_content.index('tcp')
            udp_index = cell_content.index('udp')
            if tcp_index > udp_index:
                tcp_ports, err_msg = _extractNumber(cell_content[tcp_index + 3:])
                error_msg += err_msg
                udp_ports, err_msg = _extractNumber(cell_content[:tcp_index])
                error_msg += err_msg
            else:
                tcp_ports, err_msg = _extractNumber(cell_content[:udp_index])
                error_msg += err_msg
                udp_ports, err_msg = _extractNumber(cell_content[udp_index + 3:])
                error_msg += err_msg
    elif tcp:
        if 'udp' in cell_content:
            error_msg += '只选择了TCP但是出现UDP端口; '
        else:
            tcp_ports, err_msg = _extractNumber(cell_content)
            error_msg += err_msg
    elif udp:
        if 'tcp' in cell_content:
            error_msg += '只选择了UDP但是出现TCP端口; '
        else:
            udp_ports, err_msg = _extractNumber(cell_content)
            error_msg += err_msg
    else:
        error_msg += '未选择传输协议; '
    # print({'tcp': tcp_ports, 'udp': udp_ports}, error_msg)
    if not re.findall(r'\d+', cell_content):
        error_msg += '[{}]没有提取到端口信息; '.format(cell_content)
    return {'tcp': sorted(tcp_ports), 'udp': sorted(udp_ports)}, error_msg


def _completePortNumber(port_from, port_to):
    '''
    补全端口“从”-“到”之间的连续数字，确保tcp端口已经按升序排序
    :param port_from: {'tcp': 'xx\nxx', 'udp': 'xx\nxx'}
    :param port_to: {'tcp': 'xx\nxx', 'udp': 'xx\nxx'}
    :return: {'tcp': 'xx\nxx', 'udp': 'xx\nxx'}
    '''
    f_tcp = port_from['tcp']
    f_udp = port_from['udp']
    t_tcp = port_to['tcp']
    t_udp = port_to['udp']
    result = {'tcp': '', 'udp': ''}
    error_msg = ''
    if not ((f_tcp and t_tcp) or (f_udp and t_udp)):
        error_msg += '端口【从】和【到】协议不相同/'
    elif f_tcp and t_tcp:
        ports, err_msg = _checkPort(f_tcp, t_tcp)
        if err_msg:
            error_msg += err_msg
        else:
            result['tcp'] = '\n'.join(str(i) for i in ports)
    elif f_udp and t_udp:
        ports, err_msg = _checkPort(f_tcp, t_tcp)
        if err_msg:
            error_msg += err_msg
        else:
            result['udp'] = '\n'.join(str(i) for i in ports)
    if error_msg:
        error_msg = error_msg.strip('/') + '; '
    return result, error_msg


def _checkPort(f_ports, t_ports):
    '''

    :param f_ports:
    :param t_ports:
    :return:
    '''
    ports = []
    error_msg = ''
    if not len(f_ports.split('\n')) == len(t_ports.split('\n')):
        error_msg += '端口【从】和【到】端口号数量不同;'
        return ports, error_msg

    f_ports = f_ports.split('\n')
    t_ports = t_ports.split('\n')
    for i in range(len(f_ports)):
        if f_ports[i] > t_ports[i]:
            error_msg += '端口【从】大于【到】端口号数值/'
            break
        elif (i + 1) > len(f_ports):
            if t_ports[i] > f_ports[i + 1]:
                error_msg += '端口【从】和【到】端口号数值范围存在重叠/'
                break
    if error_msg:
        return ports, error_msg

    items = zip(f_ports, t_ports)
    for s, e in items:
        if s > e:
            error_msg += '端口【从】大于【到】端口号数值/'
        else:
            ports.extend(_completeNumber(s + '-' + e))
    return ports, error_msg


def _sortItem(item_dict):
    '''
    将每一行的字典值转化为列表值，以便存入到数据库中。前提是前面的已经将错误信息处理了
    :param item_dict:
    :return:['direction', 'source_IP', 'source_map_IP', 'source_port', 'dest_IP', 'dest_map_IP', 'dest_port',
            'transport_protocal', 'app_protocal', 'access_use', 'vpn_domain']
    '''
    error_msg = ''
    items = []
    # 处理源端口信息
    s_port_from = item_dict['source_port_from']
    s_port_to = item_dict['source_port_to']
    if s_port_from and s_port_to:
        if s_port_from == s_port_to:
            item_dict['source_port'] = s_port_from
        else:
            ports, err_msg = _completePortNumber(s_port_from, s_port_to)
            if err_msg:
                error_msg += err_msg
            else:
                item_dict['source_port'] = ports
    elif not (s_port_from or s_port_to):
        item_dict['source_port'] = '不限'
    elif s_port_from:
        item_dict['source_port'] = s_port_from
    else:
        item_dict['source_port'] = s_port_to
    # 处理目的端口信息
    d_port_from = item_dict['dest_port_from']
    d_port_to = item_dict['dest_port_to']
    if d_port_from and d_port_to:
        if d_port_from == d_port_to:
            item_dict['dest_port'] = d_port_from
        else:
            ports, err_msg = _completePortNumber(d_port_from, d_port_to)
            if err_msg:
                error_msg += err_msg
            else:
                item_dict['dest_port'] = ports
    elif not (d_port_from or d_port_to):
        error_msg = '没有指定目的端口号; '
    elif d_port_from:
        item_dict['dest_port'] = d_port_from
    else:
        item_dict['dest_port'] = d_port_to
    # 前面出现错误信息则终止
    if error_msg: return items, error_msg
    # 拆分同时含有TCP和UDP协议的端口为2条信息
    if item_dict['dest_port']['tcp']:
        items.append(
            [item_dict['direction'], item_dict['source_IP'], item_dict['source_map_IP'], item_dict['source_port'],
             item_dict['dest_IP'], item_dict['dest_map_IP'], item_dict['dest_port']['tcp'], 'tcp',
             item_dict['app_protocal'], item_dict['access_use'], item_dict['vpn_domain']])
    if item_dict['dest_port']['udp']:
        items.append(
            [item_dict['direction'], item_dict['source_IP'], item_dict['source_map_IP'], item_dict['source_port'],
             item_dict['dest_IP'], item_dict['dest_map_IP'], item_dict['dest_port']['udp'], 'udp',
             item_dict['app_protocal'], item_dict['access_use'], item_dict['vpn_domain']])

    return items, error_msg


def _writeContent(tableName, filename, data):
    '''
    生成网络策略开通文件
    :param tableName:
    :param filename:
    :param data:
    :return:
    '''
    file_pre, ext = os.path.splitext(filename)
    filename = '{0}{1}{2}'.format(file_pre[:-20], file_pre[-20:-9], ext)  # 去掉自动添加上去的时间戳
    file_item = filename.split(os.sep)
    file_item[-2] = 'download'
    filename = os.sep.join((file_item))
    if os.path.exists(filename): os.remove(filename)  # 删除已经存在的文件
    wb = Workbook()
    for k, v in data.items():
        ws = wb.create_sheet(k, 0)
        for row in range(len(v)):
            for col in range(len(v[row])):
                ws.cell(row=row + 1, column=col + 1, value=v[row][col])  # cell 起始位置必须是1而非0
    wb.save(filename)
    wb.close()
    cache.set('download{0}{1}file'.format('network', tableName), filename, 1 * 60)


def importIPMappingXls(tableName, filename):
    '''
    读取表格数据并入库
    :param tableName:
    :param filename:
    :return: 入库结果和返回的HTML错误信息
    '''
    print('处理CZW IP文件：{}'.format(filename))
    datas = defaultdict(dict)
    message = ''
    warning_msgs = []
    error_msgs = []
    wb = load_workbook(filename, data_only=True)
    sheet_list = wb.sheetnames
    for sheet_name in sheet_list:
        data = []
        errors = []  # 存储检查本表单发现的错误信息
        warnings = []  # 存储检查本表单发现的异常信息
        province = ''
        province_flag = False  # 判断表单是否含有省份名称
        pat_flag = False  # 判断表单是否为PAT表单
        head_flag = False  # 判断是否为表单表头
        sheet = wb[sheet_name]
        sheet_name = sheet.title
        # 判断表名是否为省公司名称
        for pl in province_list:
            if pl in sheet_name:
                province = pl
                province_flag = True
        if not province_flag:
            print('表单【{}】的省份名字出现问题，请检查后再导入'.format(sheet_name))
            error_msgs.append('表单【{}】的省份名字出现问题，请检查后再导入'.format(sheet_name))
            return message, datas

        # 判断表是否为PAT/NAT表
        if 'pat' in sheet_name.lower():
            pat_flag = True
        # 处理表单内容
        # 拆分并填充合并单元格
        sheet = _unmergeCell(sheet)
        for row in sheet.rows:
            cell_values = [_cellValue(c) for c in row]
            # 处理表单第一步：需要确认表单表头内容，但不记录表头元素所在的列
            if '地址' in '|'.join(cell_values):  # 排除第一行
                head_flag = True
                continue
            if not head_flag:  # 没有找到表单表头则跳过
                continue
            # 处理单行内容
            # 判断IP地址是否合法，格式化元素，返回信息：
            #       PAT表：  错误信息，[源地址，源端口，目的地址，目的端口，类型，网络号，'']
            #       非PAT表： 错误信息，[源地址，目的地址，系统]
            error_msg, warning_msg, items = _mappingRowItems(cell_values, pat_flag)
            if error_msg:
                errors.append(error_msg)
                continue
            if warning_msg: warnings.append(warning_msg)
            data.append(items)
        if pat_flag:
            datas[province]['PAT'] = data
        else:
            datas[province]['SINGLE'] = data
        if errors: error_msgs.append('处理表单【{}】存在非法错误：\n\t{}'.format(sheet_name, '\n\t'.join(errors)))
        if warnings: warning_msgs.append('处理表单【{}】时，发现需要注意的事项：\n\t{}'.format(sheet_name, '\n\t'.join(warnings)))
    for province, data in datas.items():
        errors = []
        warnings = []
        s_error, s_data = _provinceContinueIPComplete(data['SINGLE'])
        if s_error: errors.append(s_error)
        p_error, p_warning, s_data, p_data = _provinceItemInfoCompelte(s_data, data['PAT'])
        if p_error: errors.append(p_error)
        if p_warning: warnings.append(p_warning)
        if errors: error_msgs.append('处理省份【{}】省份存在错误\n\t{}'.format(province, '\n\t'.join(errors)))
        if warnings: warning_msgs.append('处理省份【{}】时，发现需要注意的事项\n\t{}'.format(province, '\n\t'.join(warnings)))
        # 更新最终结果
        datas[province]['SINGLE'] = s_data
        datas[province]['PAT'] = p_data

    db_result = _dbOps(datas)
    if warning_msgs:
        message += '<div style="background:#FF0">{}</div>'.format(
            ''.join(['<p>{}</p>'.format(i) for i in warning_msgs]).replace('\t', '&emsp;' * 2).replace('\n', '</p><p>'))
    if error_msgs:
        message += '<div style="color:#F00">{}</div>'.format(
            ''.join(['<p>{}</p>'.format(i) for i in error_msgs]).replace('\t', 'emsp;' * 2).replace('\n', '</p><p>'))

    return db_result, message


def _dbOps(datas):
    '''
    数据入库操作
    :param datas:{province:{'SINGLE':[], 'PAT':[]}, ...}
    :return:
    '''
    m_import_result, m_update_result, p_import_result, p_update_result = True, True, True, True

    ip_mapping = []
    ip_pat_mapping = []
    ip_db_col = tableColums['ipmapping']
    pat_db_col = tableColums['ippatmapping']
    # 组织数据
    for province, v in datas.items():
        pat = v['PAT']
        for item in pat:
            item.insert(0, province)
            ip_pat_mapping.append(dict(zip(pat_db_col, item)))
        single = v['SINGLE']
        for item in single:
            item.insert(0, province)
            ip_mapping.append(dict(zip(ip_db_col, item)))

    db_ip_mapping = getAll('ipmapping')
    db_ip_pat_mapping = getAll('ippatmapping')
    if not db_ip_mapping:
        print('import to ipmapping table')
        m_import_result = importDatabase('ipmapping', ip_mapping)
    else:
        print('update to ipmapping table')
        m_update_result = updateBulk('ipmapping', ip_mapping, 'source_ip')

    if not db_ip_pat_mapping:
        print('import to ippatmapping table')
        p_import_result = importDatabase('ippatmapping', ip_pat_mapping)
    else:
        print('update to ippatmapping table')
        p_update_result = updateBulk('ippatmapping', ip_pat_mapping,
                                     ['source_ip', 'source_port', 'dest_ip', 'dest_port'])

    for i in [m_import_result, m_update_result, p_import_result, p_update_result]:
        if not i:
            return False
    return True


def _patContinueIPComplete(ip_item, split_tag='-'):
    '''
    补全连续IP地址
    :param ip_item:
    :param split_tag:
    :return:
    '''
    result = []
    start_ip, end_ip = ip_item.split(split_tag)
    ip_networks = [ipaddr for ipaddr in summarize_address_range(IPv4Address(start_ip), IPv4Address(end_ip))]
    for ip_network in ip_networks:
        result.extend([str(ipaddr) for ipaddr in IPv4Network(ip_network)])
    return result


def _provinceItemInfoCompelte(s_data, p_data):
    '''
    根据PAT表填充非PAT表，合并PAT表中已经存在的项
    :param s_data: 非PAT表： [源地址，目的地址，系统]
    :param p_data: PAT表：   [源地址，源端口，目的地址，目的端口，类型，网络号，'']
    :return:    s_data, p_data:
    '''
    errors = []
    warnings = []
    p_dict = {}
    contained_ip = {}  # 存储网络号所含的IP地址

    s_ip_list = [ip[0] for ip in s_data]
    p_ip_list = [ip[0] for ip in p_data]
    for pil in set(p_ip_list):
        ip_list = []
        system = ''
        # 处理连续PAT地址
        if '-' in pil:
            ip_list = _patContinueIPComplete(pil)
        else:
            ip_list.append(pil)
        for il in ip_list:
            if not il in s_ip_list:
                errors.append('{}的PAT源地址不存在'.format(pil))
                break
            # 填充PAT信息
            map_ip = s_data[s_ip_list.index(il)][1]
            if map_ip == 'PAT':  # 分类正确，不处理
                continue
            elif map_ip:  # 分类被占用
                warnings.append('{} 的PAT映射地址被 {} 占用'.format(il, map_ip))
            else:  # 未分类，填充分类内容
                s_data[s_ip_list.index(il)][1] = 'PAT'
            # 填充PAT表系统信息
            if not system:
                system = s_data[s_ip_list.index(il)][2]
            # 连续地址对应的系统名称不一致
            elif not system == s_data[s_ip_list.index(il)][2]:
                warnings.append('连续PAT IP地址中存在对应不同系统的情况')
        for index, item in enumerate(p_data):
            if pil == item[0]:
                p_data[index][-1] = system
    # 以源IP地址为键，建立字典
    for item in p_data:
        if item[0] in p_dict:
            p_dict[item[0]].append(item[1:])
        else:
            p_dict[item[0]] = [item[1:]]
    for src_ip, values in p_dict.items():
        # 提取包含多个IP地址的网络号
        net_works = [ip[-2] for ip in values if '/' in ip[-2] and '/32' not in ip[-2]]
        if not net_works: continue
        ip_list = sorted([ip[1] for ip in values], key=socket.inet_aton)
        # 判断地址是否在包含多地址的网络号中
        for net_work in net_works:
            for ip in ip_list:
                if IPv4Address(ip) in IPv4Network(net_work):
                    # 找出非网络号的IP地址
                    if not ip == net_work.split('/')[0]:
                        if not net_work in contained_ip:
                            contained_ip[net_work] = [(ip, src_ip), ]
                        else:
                            contained_ip[net_work].append((ip, src_ip))
    # 删除被包含的IP地址信息
    dup_items = []
    for k, v in contained_ip.items():
        warnings.append('{} 被包含在 {} 中'.format('、'.join([i[0] for i in v]), k))
        for ip in v:
            dup_items.extend([item for item in p_data if item[0] == ip[1] and item[2] == ip[0]])
    [p_data.remove(di) for di in dup_items]
    # 去除PAT源地址和映射地址组合值为重复的
    combine_list = [(item[0], item[1], item[2], item[3]) for item in p_data]
    duplicate_ip = [item for item, count in Counter(combine_list).items() if count > 1]
    if duplicate_ip:
        for di in duplicate_ip:
            warnings.append('源地址{}源端口{}映射地址{}映射端口{}'.format(di[0], di[1], di[2], di[3]))

    return '；\n\t'.join(errors), '；\n\t'.join(warnings), s_data, p_data


def _provinceContinueIPComplete(ip_data):
    '''
    根据现有省公司地址，提取前24位网络地址，并补全后面的IP地址
    不能对PAT表的ip地址去重
    :param province_ip_data: 含源地址、映射地址、系统名称的信息 [[src_ip, map_ip, system]...]
    :return:
    '''
    result = []
    error = ''
    dips = []  # 重复IP地址
    # 找到网络号
    ip_networks = set(['.'.join(ip[0].split('.')[:3]) + '.0/24' for ip in ip_data])
    ip_list = [d[0] for d in ip_data]
    duplicate_ip = [item for item, count in Counter(ip_list).items() if count > 1]
    if duplicate_ip:
        for dip in duplicate_ip:
            if ip_data[ip_list.index(dip)][1]:
                dips.append(dip)
    if dips:
        error = error + '存在重复IP地址{}'.format(','.join(dips))
        return error, result

    for ip_network in ip_networks:
        for ip in IPv4Network(ip_network):
            ip = str(ip)
            if ip in ip_list:
                result.append(ip_data[ip_list.index(ip)])
            else:
                result.append([ip, '', ''])
    return error, result


def _IPPortCheck(port):
    '''
    判断端口号是否合法
    :param port:
    :return:
    '''
    if not (isinstance(port, int) or isinstance(port, str)):
        return False
    if isinstance(port, str):
        port = port.strip()
    if not port:
        return False
    try:
        port = int(port)
    except:
        return False
    if port < 1 or port > 65535:
        return False
    return True


def _mappingRowItems(cell_values, pat_flag):
    '''
    判断一行数据中的IP地址/IP端口是否合法，若是PAT还需生成映射地址网络号xxx.xxx.xxx.xxx/xx
    :param cell_values:
    :param pat_flag:
    :return: PAT表：   错误信息，[源地址，源端口，目的地址，目的端口，类型，网络号，'']
             非PAT表： 错误信息，[源地址，目的地址，系统]
    '''
    errors = []
    warnings = []
    src_ip = cell_values[0]
    src_port = ''
    map_port = ''
    ip_network = ''
    type = ''
    system = ''
    if pat_flag:
        src_port = cell_values[1]
        map_ip = cell_values[2]
        map_port = cell_values[3]
        type = cell_values[4].upper()
    else:
        map_ip = cell_values[1]
        system = cell_values[2]
    if not src_ip:
        return '源地址为空', warnings, ['', '', '']
    if '-' in src_ip:
        start_ip, end_ip = src_ip.split('-')
        if not _isIp(start_ip) or not _isIp(end_ip):
            errors.append('源地址 {} 不合法'.format(src_ip))
    elif (not _isIp(src_ip)) or src_ip.startswith('255.'):
        errors.append('源地址 {} 不合法'.format(src_ip))

    if pat_flag:  # PAT表单处理
        if (not _isIp(map_ip)) or map_ip.startswith('255.'):
            errors.append('映射地址 {} 不合法'.format(map_ip))
        if src_port and not _IPPortCheck(src_port):
            errors.append('源端口号 {} 不合法'.format(src_port))
        if map_port:
            if map_port.startswith('255.'):  # 计算网络号
                ip_network = str(IPv4Network('{}/{}'.format(cell_values[2], cell_values[3])))
                if not type == 'PAT': warnings.append('{} 类型应为PAT'.format(map_ip))
                if int(ip_network.split('/')[1]) <= 24: warnings.append('单一地址 {} 分配过多映射地址'.format(map_ip))
                type = 'PAT'
            elif not _IPPortCheck(map_port):
                errors.append('映射端口号 {} 不合法'.format(map_port))
            else:
                if not type == 'NAT': warnings.append('{} 类型应为NAT'.format(map_ip))
                type = 'NAT'
        else:
            errors.append('映射端口号 {} 不能为空'.format(map_port))
    else:  # 普通表单处理
        if map_ip and not map_ip.upper() == 'PAT' and (
                not _isIp(map_ip) or map_ip.startswith('255.') or map_ip.endswith('.0')):
            errors.append('映射地址 {} 不合法'.format(map_ip))
        # 网络号不允许分配映射地址
        if src_ip.endswith('.0') and map_ip:
            errors.append('源地址为网络号，但存在映射地址 {} '.format(map_ip))
    if errors:  # 格式化错误信息
        errors = 'IP地址/端口[ {} ]'.format(', '.join(errors))
    if warnings:  # 格式化错误信息
        warnings = 'IP地址/端口[ {} ]'.format(', '.join(warnings))
    if pat_flag:
        return errors, warnings, [src_ip, src_port, map_ip, map_port, type, ip_network, '']
    return errors, warnings, [src_ip, map_ip, system]


def _unmergeCell(sheet_content):
    '''
    拆分并填充合并单元格
    :param sheet_content:
    :return:
    '''
    m_cells = sheet_content.merged_cells.ranges
    m_cells_dict = {}
    for mc in m_cells:
        cell_value = sheet_content.cell(row=mc.min_row, column=mc.min_col).value
        m_cells_dict[mc.coord] = [mc, cell_value]
    for k, v in m_cells_dict.items():
        merge_all_list = []
        sheet_content.unmerge_cells(k)
        r1, r2, c1, c2 = v[0].min_row, v[0].max_row, v[0].min_col, v[0].max_col
        if (r1 != r2 and c1 != c2):
            row_col = [(x, y) for x in range(r1, r2 + 1) for y in range(c1, c2 + 1)]
            merge_all_list.extend(row_col)
        elif (r1 == r2 and c1 != c2):  # or (r1 != r2 and c1 == c2):
            col = [(r1, n) for n in range(c1, c2 + 1)]
            merge_all_list.extend(col)
        elif (r1 != r2 and c1 == c2):
            row = [(m, c1) for m in range(r1, r2 + 1)]
            merge_all_list.extend(row)
        for mal in merge_all_list:
            sheet_content.cell(row=mal[0], column=mal[1]).value = v[1]

    return sheet_content


def complateAccessList(device, originalIP, mask, distinateIP, port):
    data = ''
    errors = {'originalIP': '', 'mask': '', 'distinateIP': '', 'port': ''}
    originalIP = originalIP.split()
    distinateIP = distinateIP.split()

    if not _isIp(mask):
        errors['mask'] = mask
        return data, errors

    if len(port.split()) > 1:
        port_str = ' range '
    else:
        port_str = ' eq '

    for o in originalIP:
        if not _isIp(o):
            if errors['originalIP']:
                errors['originalIP'] = errors['originalIP'] + ',' + o
            else:
                errors['originalIP'] = o
            continue
        for d in distinateIP:
            if not _isIp(d):
                if errors['distinateIP']:
                    errors['distinateIP'] = errors['originalIP'] + ',' + d
                else:
                    errors['distinateIP'] = d
                continue
            if device == 'CISCO':
                accesslist_str = 'access-list outside extended permit tcp ' + str(o) + ' ' + str(mask) + ' host ' + str(
                    d) + port_str + ' ' + str(port)
            elif device in ['H3C', 'huawei']:
                accesslist_str = 'rule permit tcp source ' + str(o) + ' ' + str(mask) + ' destination ' + str(
                    d) + port_str + ' ' + str(port)
            else:
                accesslist_str = '请选择厂商'
            data = data + '<p>' + accesslist_str + '</p>'
    return data, errors

# importIPMappingXls('', r'D:\Documents\2Desktop\基础平台部\04 网络日常工作\承载网1105.xlsx')
